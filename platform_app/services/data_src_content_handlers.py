"""获取数据后的内容处理策略（可扩展框架）。

根据 DataSrc.fetch_mode 决定：保存原始响应，或从 HTML 中提取表格并按 data_src.format_type 保存。
- RAW: 原样保存响应字节。
- HTML_TABLES: 解析 HTML 提取所有 <table>，按 format_type 保存为 csv / xlsx / xls。
"""
import csv
import io
import logging
from typing import List, Tuple

from platform_app.models import FormatType

logger = logging.getLogger(__name__)


def _extract_tables_from_html(html_bytes: bytes) -> List[List[List[str]]]:
    """从 HTML 字节中解析所有 <table>，每个 table 转为二维列表 [行][列]。"""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise RuntimeError("请安装 beautifulsoup4：pip install beautifulsoup4")

    soup = BeautifulSoup(html_bytes, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return []

    result = []
    for tbl in tables:
        rows = []
        trs = tbl.find_all("tr")
        for tr in trs:
            cells = []
            for cell in tr.find_all(["td", "th"]):
                cells.append(cell.get_text(strip=True))
            if cells:
                rows.append(cells)
        if rows:
            result.append(rows)
    return result


def _tables_to_xlsx_bytes(tables: List[List[List[str]]]) -> bytes:
    """将多个二维表写入一个 xlsx（每表一个 sheet），返回文件字节。"""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("请安装 openpyxl：pip install openpyxl")

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for idx, rows in enumerate(tables):
        sheet_name = f"Table{idx + 1}"[:31]
        ws = wb.create_sheet(sheet_name, index=idx)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _tables_to_xls_bytes(tables: List[List[List[str]]]) -> bytes:
    """将多个二维表写入一个 .xls（每表一个 sheet），返回文件字节。"""
    try:
        import xlwt
    except ImportError:
        raise RuntimeError("请安装 xlwt：pip install xlwt")

    buf = io.BytesIO()
    wb = xlwt.Workbook()
    for idx, rows in enumerate(tables):
        sheet_name = f"Table{idx + 1}"[:31]
        ws = wb.add_sheet(sheet_name)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _tables_to_csv_bytes(tables: List[List[List[str]]]) -> bytes:
    """将第一个表写入 CSV。"""
    if not tables:
        return b""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in tables[0]:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def _tables_to_json_bytes(tables: List[List[List[str]]]) -> bytes:
    """将第一个表转为 JSON 数组（首行作 key）。"""
    import json
    if not tables or not tables[0]:
        return b"[]"
    rows = tables[0]
    keys = rows[0]
    out = [dict(zip(keys, row)) for row in rows[1:]]
    return json.dumps(out, ensure_ascii=False).encode("utf-8")


def _tables_to_format_bytes(tables: List[List[List[str]]], format_type: int) -> Tuple[bytes, str]:
    """按 format_type 将 tables 转为 (字节, 扩展名)。"""
    if format_type == FormatType.CSV:
        return _tables_to_csv_bytes(tables), ".csv"
    if format_type == FormatType.XLS:
        return _tables_to_xls_bytes(tables), ".xls"
    if format_type == FormatType.JSON:
        return _tables_to_json_bytes(tables), ".json"
    # FormatType.EXCEL 或默认
    return _tables_to_xlsx_bytes(tables), ".xlsx"


def process_fetched_content(
    fetch_mode: int,
    raw_bytes: bytes,
    resolved_url: str,
    output_format_type: int,
) -> Tuple[bytes, str, int]:
    """
    根据 fetch_mode 处理拉取到的原始字节，返回 (要写入文件的字节, 建议扩展名, format_type)。

    - fetch_mode FetchMode.RAW: 原样返回，扩展名从 URL 推断。
    - fetch_mode FetchMode.HTML_TABLES: 解析 HTML 提取 <table>，按 output_format_type 保存为 csv/xlsx/xls。
    """
    from platform_app.models import FetchMode

    if fetch_mode == FetchMode.HTML_TABLES:
        tables = _extract_tables_from_html(raw_bytes)
        if not tables:
            raise ValueError("页面中未找到任何 <table>，请确认 URL 返回的是包含表格的 HTML，或改用「原始」模式")
        out_bytes, ext = _tables_to_format_bytes(tables, output_format_type)
        return out_bytes, ext, output_format_type

    # RAW：原样
    ext = ""
    lower = (resolved_url or "").lower()
    if ".xlsx" in lower or ".xls" in lower:
        ext = ".xlsx" if ".xlsx" in lower else ".xls"
    elif ".csv" in lower:
        ext = ".csv"
    elif ".json" in lower:
        ext = ".json"
    format_type = FormatType.JSON
    if ext in (".xlsx", ".xls"):
        format_type = FormatType.EXCEL if ext == ".xlsx" else FormatType.XLS
    elif ext == ".csv":
        format_type = FormatType.CSV
    return raw_bytes, ext, format_type
